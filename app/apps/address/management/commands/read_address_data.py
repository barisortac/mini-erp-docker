# -*- coding: utf-8 -*-
from __future__ import unicode_literals

"""
http://postakodu.ptt.gov.tr/  : Turkey's city, township, districts
"""

import os
import openpyxl

from django.core.management.base import BaseCommand

from address.models import Country, City
from address.get_countries import RestCountryApi


class Command(BaseCommand):
    help = 'Inserts city, township and districts values'

    def title(self, val):
        val = val.replace(u'I', 'X')
        val = val.replace('İ', 'i')
        val = val.title()
        val = val.replace(u'x', u'ı').replace(u'X', u'I')
        return val

    def read(self, path, sheet):
        normalize = {}

        for x in sheet.iter_rows(min_row=2):
            # row = list(sheet.iter_rows())[x]
            # excel file columns
            # city, township, district, neighborhood, postal_code
            try:
                c, t, d, n = x
            except Exception:
                c, t, d, n, p = x

            c = self.title(c.value.strip())
            t = self.title(t.value.strip())
            d = self.title(d.value.strip())
            n = self.title(n.value.strip())

            nil = {}
            if c in normalize:
                nil[c] = normalize[c]
            else:
                nil[c] = {}

            nilce = {}
            if t in nil[c]:
                nilce[t] = nil[c][t]
            else:
                nilce[t] = []

            nilce[t].append(n)

            nil[c].update(nilce)
            normalize.update(nil)

        return normalize

    def insert(self, path):
        # excel = xlrd.open_workbook(path, encoding_override="utf-8")
        # sheet = excel.sheet_by_index(0)
        wb = openpyxl.load_workbook(path, read_only=True)
        sheet = wb.active
        data = self.read(path, sheet)

        country, cr = Country.objects.get_or_create(name="Türkiye")

        for d in data:
            city, x = City.objects.get_or_create(name=d, country=country)
            print("city --> ", d)
            # for t in data[d]:
            # print "township --> ", t
            # township, x = Township.objects.get_or_create(name=t, city=city)
            # for di in data[d][t]:
            # print "district --> ", di
            # district, x = District.objects.get_or_create(name=di, township=township)

        return True

    def handle(self, *args, **kwargs):
        rca = RestCountryApi()
        rca.insert()

        base_path = os.path.dirname(os.path.abspath(__file__))
        path = base_path + "/pk_list.xlsx"

        self.insert(path)
